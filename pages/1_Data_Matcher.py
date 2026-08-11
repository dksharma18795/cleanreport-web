import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import io
import re
import difflib
import docx
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# --- WEB APP UI SETUP ---
# ==========================================
st.set_page_config(
    page_title="CleanReport - Ultimate Data Matcher", 
    page_icon="⚡", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

ga_code = """
<!-- Google tag (gtag.js) -->
<script async src="https://www.googletagmanager.com/gtag/js?id=G-VEH0V9QEEV"></script>
<script>
  window.dataLayer = window.dataLayer || [];
  function gtag(){dataLayer.push(arguments);}
  gtag('js', new Date());

  gtag('config', 'G-VEH0V9QEEV');
</script>
"""
components.html(ga_code, height=0, width=0)

hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;} 
            footer {visibility: hidden;}    
            [data-testid="collapsedControl"] {display: none;}
            [data-testid="stSidebarNav"] {display: none;} 
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)

# ==========================================
# --- NAVIGATION ---
# ==========================================
if st.button("🏠 Back to Dashboard"):
    st.switch_page("app.py")

# ==========================================
# --- CORE MATCHING ENGINE ---
# ==========================================
def get_pure_core(name):
    if not isinstance(name, str) or str(name).lower() == 'nan': return ""
    n = name.upper()
    n = n.replace("M/S.", " ").replace("M/S", " ").replace("MS.", " ")
    n = re.sub(r'\b(PVT|PRIVATE|LTD|LIMITED|P LTD|CO|CORP|CORPORATION|INC|AND|LLP|INDIA|THE|ENTERPRISES|INDUSTRIES|PRODUCTS|COMPANY)\b', ' ', n)
    return re.sub(r'[^A-Z0-9]', '', n)

def standardize_company(raw_name, master_standards, master_core_dict, master_core_list):
    if not raw_name or str(raw_name).strip() == "": return None
    orig_str = str(raw_name).strip().upper()
    for m in master_standards:
        if orig_str == m.upper(): return m
    core_orig = get_pure_core(orig_str)
    if not core_orig: return None
    if core_orig in master_core_dict: return master_core_dict[core_orig]
    for m_core, m_orig in master_core_dict.items():
        if len(m_core) >= 5 and len(core_orig) >= 5:
            if m_core in core_orig or core_orig in m_core:
                if difflib.SequenceMatcher(None, core_orig, m_core).ratio() > 0.6:
                    return m_orig
    matches = difflib.get_close_matches(core_orig, master_core_list, n=1, cutoff=0.75)
    if matches: return master_core_dict[matches[0]]
    return None

def extract_text_from_file(uploaded_file):
    text = ""
    try:
        if uploaded_file.name.endswith(('.xlsx', '.xls')):
            engine = 'calamine' if uploaded_file.name.endswith('.xls') else 'openpyxl'
            df = pd.read_excel(uploaded_file, header=None, engine=engine)
            text = "\n".join(df.iloc[:, 0].dropna().astype(str).tolist())
        elif uploaded_file.name.endswith('.docx'):
            doc = docx.Document(uploaded_file)
            text = "\n".join([p.text for p in doc.paragraphs])
        elif uploaded_file.name.endswith('.txt'):
            text = uploaded_file.read().decode('utf-8')
    except Exception as e:
        st.error(f"File reading error: {e}")
    return text

def read_master_file(uploaded_file):
    engine = 'calamine' if uploaded_file.name.endswith('.xls') else 'openpyxl'
    return pd.read_excel(uploaded_file, engine=engine)

# ==========================================
# --- MAIN AREA ---
# ==========================================
st.title("⚡ CleanReport: Ultimate Excel Data Matcher")
st.markdown("---")

st.subheader("⚙️ 1. Upload Master Files")
c1, c2 = st.columns(2)
with c1:
    sales_file = st.file_uploader("1. Upload Sales File", type=["xlsx", "xls"])
with c2:
    stock_file = st.file_uploader("2. Upload Stock File", type=["xlsx", "xls"])

st.markdown("<br>", unsafe_allow_html=True)
st.subheader("🎯 2. Upload Target Companies list file")
col1, col2 = st.columns(2)
with col1:
    target_file = st.file_uploader("Option A: Upload List (Excel, Word, TXT)", type=["xlsx", "xls", "docx", "txt"])
with col2:
    raw_text_input = st.text_area("Option B: Paste Raw Names (One name per line)", height=150)

st.markdown("<br>", unsafe_allow_html=True)

if st.button("🚀 Process & Generate Clean Report", use_container_width=True, type="primary"):
    if sales_file and stock_file and (target_file or raw_text_input.strip()):
        with st.spinner("Processing data, applying professional formatting..."):
            try:
                # Load Master Data
                df_sales = read_master_file(sales_file)
                df_stock = read_master_file(stock_file)
                
                sales_companies_series = df_sales.iloc[:, 1].astype(str).str.strip().str.upper()
                stock_companies_series = df_stock.iloc[:, 0].astype(str).str.strip().str.upper()
                
                sales_unique_raw = set(sales_companies_series.unique())
                stock_unique_raw = set(stock_companies_series.unique())
                
                master_standards = list(set(list(sales_unique_raw) + list(stock_unique_raw)))
                master_standards = [str(x).strip() for x in master_standards if str(x).strip() != 'NAN']
                
                master_core_dict = {}
                for m_name in master_standards:
                    core = get_pure_core(m_name)
                    if core and core not in master_core_dict:
                        master_core_dict[core] = m_name
                master_core_list = list(master_core_dict.keys())

                # Combine text
                combined_text = ""
                if target_file:
                    combined_text += extract_text_from_file(target_file) + "\n"
                if raw_text_input:
                    combined_text += raw_text_input
                
                raw_names_list = [name.strip() for name in combined_text.split('\n') if name.strip()]
                total_provided = len(raw_names_list)
                
                valid_target_names = set()
                not_found_names = []

                for raw_name in raw_names_list:
                    standard_name = standardize_company(raw_name, master_standards, master_core_dict, master_core_list)
                    if standard_name:
                        valid_target_names.add(standard_name)
                    else:
                        not_found_names.append(raw_name)

                found_in_sales = [c for c in valid_target_names if c in sales_unique_raw]
                found_in_stock = [c for c in valid_target_names if c in stock_unique_raw]
                missing_in_sales = [c for c in valid_target_names if c not in sales_unique_raw]
                missing_in_stock = [c for c in valid_target_names if c not in stock_unique_raw]
                
                # Filter Data
                sales_filtered = df_sales[sales_companies_series.isin(valid_target_names)].copy()
                stock_filtered = df_stock[stock_companies_series.isin(valid_target_names)].copy()
                
                columns_to_drop = ['INDENT_NUM', 'STATE_NAME', 'HSN_CODE']
                sales_filtered = sales_filtered.drop(columns=[col for col in columns_to_drop if col in sales_filtered.columns])
                
                # SORTING
                if len(sales_filtered.columns) > 1:
                    sales_filtered = sales_filtered.sort_values(by=sales_filtered.columns[1], ascending=True)
                if len(stock_filtered.columns) > 0:
                    stock_filtered = stock_filtered.sort_values(by=stock_filtered.columns[0], ascending=True)
                
                # Stats Table
                df_stats = pd.DataFrame({
                    "DATA ANALYTICS METRIC": [
                        "Total Companies Provided (Raw Input)",
                        "Successfully Standardized (Found in Masters)",
                        "Companies Present in SALES File",
                        "Companies Present in STOCK File",
                        "Missing from SALES File",
                        "Missing from STOCK File",
                        "NOT FOUND in Any Master File"
                    ],
                    "COUNT": [
                        total_provided, len(valid_target_names), len(found_in_sales),
                        len(found_in_stock), len(missing_in_sales), len(missing_in_stock), len(not_found_names)
                    ]
                })

                max_len = max(len(missing_in_sales), len(missing_in_stock), len(not_found_names))
                df_lists = pd.DataFrame({
                    "Missing in Sales (List)": missing_in_sales + [""] * (max_len - len(missing_in_sales)),
                    "Missing in Stock (List)": missing_in_stock + [""] * (max_len - len(missing_in_stock)),
                    "Not Found in Any File (List)": not_found_names + [""] * (max_len - len(not_found_names))
                })
                start_row_for_lists = len(df_stats) + 4
                
                # ==========================================
                # EXCEL 1: RAW REPORT
                # ==========================================
                output_raw = io.BytesIO()
                with pd.ExcelWriter(output_raw, engine='openpyxl') as writer:
                    sales_filtered.to_excel(writer, sheet_name="Sales_Data", index=False)
                    stock_filtered.to_excel(writer, sheet_name="Stock_Data", index=False)
                    df_stats.to_excel(writer, sheet_name="Summary_Analytics", index=False, startrow=0)
                    df_lists.to_excel(writer, sheet_name="Summary_Analytics", index=False, startrow=start_row_for_lists)
                raw_excel_data = output_raw.getvalue()
                
                # ==========================================
                # EXCEL 2: PRINTABLE CLEAN REPORT (Formatting)
                # ==========================================
                output_print = io.BytesIO()
                with pd.ExcelWriter(output_print, engine='openpyxl') as writer:
                    sales_filtered.to_excel(writer, sheet_name="Sales_Data", index=False)
                    stock_filtered.to_excel(writer, sheet_name="Stock_Data", index=False)
                    df_stats.to_excel(writer, sheet_name="Summary_Analytics", index=False, startrow=0)
                    df_lists.to_excel(writer, sheet_name="Summary_Analytics", index=False, startrow=start_row_for_lists)
                    
                    workbook = writer.book
                    
                    # --- GLOBAL STYLES ---
                    header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    alt_row_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                    white_fill = PatternFill(fill_type=None)
                    
                    thin_border = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), 
                                         top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))
                                         
                    font_12b_white = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
                    font_11b_white = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
                    
                    # --- SHEET 1: SALES DATA ---
                    ws_sales = workbook["Sales_Data"]
                    ws_sales.page_setup.orientation = ws_sales.ORIENTATION_LANDSCAPE
                    ws_sales.page_setup.paperSize = ws_sales.PAPERSIZE_A4
                    ws_sales.page_margins.left = 0.75
                    ws_sales.page_margins.right = 0.75
                    ws_sales.page_margins.top = 1.0
                    ws_sales.page_margins.bottom = 1.0
                    
                    ws_sales.row_dimensions[1].height = 96.75
                    sales_widths = {'A': 5.29, 'B': 31.29, 'C': 11.29, 'D': 16.00, 'E': 6.00, 'F': 40.00, 
                                    'G': 5.29, 'H': 5.29, 'I': 8.57, 'J': 6.43, 'K': 5.29, 'L': 8.00}
                    for col, w in sales_widths.items():
                        ws_sales.column_dimensions[col].width = w

                    # Sales Headers
                    for col_num, cell in enumerate(ws_sales[1], 1):
                        cell.fill = header_fill
                        cell.border = thin_border
                        
                        if col_num == 2:
                            cell.font = font_11b_white
                        else:
                            cell.font = font_12b_white
                            
                        if col_num in [1, 3, 4, 5, 2, 6]:
                            cell.alignment = Alignment(horizontal='center', vertical='center', textRotation=45)
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center', textRotation=45)

                    # Sales Data Rows
                    for row_num, row in enumerate(ws_sales.iter_rows(min_row=2), 2):
                        ws_sales.row_dimensions[row_num].height = 30
                        is_gray = (row_num % 2 == 0)
                        
                        for col_num, cell in enumerate(row, 1):
                            cell.fill = alt_row_fill if is_gray else white_fill
                            cell.border = thin_border
                            
                            if col_num == 1:
                                cell.font = Font(name='Calibri', size=11, bold=False, color="000000")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            elif col_num == 2:
                                cell.font = Font(name='Calibri', size=11, bold=True, color="000000")
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                            elif col_num in [3, 4, 5]:
                                cell.font = Font(name='Calibri', size=11, bold=False, color="000000")
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                            elif col_num == 6:
                                cell.font = Font(name='Calibri', size=9, bold=True, color="000000")
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                            elif col_num in [7, 8]:
                                cell.font = Font(name='Calibri', size=11, bold=True, color="166534")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            else:
                                cell.font = Font(name='Calibri', size=11, bold=False, color="000000")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # --- SHEET 2: STOCK DATA ---
                    ws_stock = workbook["Stock_Data"]
                    ws_stock.page_setup.orientation = ws_stock.ORIENTATION_LANDSCAPE
                    ws_stock.page_setup.paperSize = ws_stock.PAPERSIZE_A4
                    ws_stock.page_margins.left = 0.75
                    ws_stock.page_margins.right = 0.75
                    ws_stock.page_margins.top = 1.0
                    ws_stock.page_margins.bottom = 1.0
                    
                    ws_stock.row_dimensions[1].height = 92.25
                    stock_widths = {'A': 35.00, 'B': 6.00, 'C': 20.86, 'D': 5.29, 'E': 5.29, 'F': 5.29, 'G': 5.29, 'H': 5.29, 
                                    'I': 5.29, 'J': 5.29, 'K': 5.29, 'L': 5.29, 'M': 5.29, 'N': 5.29, 'O': 5.29, 'P': 5.29}
                    for col, w in stock_widths.items():
                        ws_stock.column_dimensions[col].width = w

                    # Stock Headers
                    for col_num, cell in enumerate(ws_stock[1], 1):
                        cell.fill = header_fill
                        cell.font = font_12b_white
                        cell.border = thin_border
                        
                        if col_num <= 3:
                            cell.alignment = Alignment(horizontal='center', vertical='center', textRotation=45)
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='top', textRotation=45)

                    # Stock Data Rows
                    for row_num, row in enumerate(ws_stock.iter_rows(min_row=2), 2):
                        ws_stock.row_dimensions[row_num].height = 36
                        is_gray = (row_num % 2 == 0)
                        
                        for col_num, cell in enumerate(row, 1):
                            cell.fill = alt_row_fill if is_gray else white_fill
                            cell.border = thin_border
                            
                            if col_num == 1:
                                cell.font = Font(name='Calibri', size=11, bold=True, color="000000")
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                            elif col_num == 2:
                                cell.font = Font(name='Calibri', size=11, bold=False, color="000000")
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                            elif col_num == 3:
                                cell.font = Font(name='Calibri', size=9, bold=True, color="000000")
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                            elif col_num == 4:
                                cell.font = Font(name='Calibri', size=11, bold=False, color="000000")
                                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                            elif col_num in [5, 6]:
                                cell.font = Font(name='Calibri', size=11, bold=True, color="000000")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            elif col_num in [7, 8]:
                                cell.font = Font(name='Calibri', size=11, bold=True, color="166534")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            elif col_num in [9, 10]:
                                cell.font = Font(name='Calibri', size=11, bold=True, color="000000")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            elif col_num in [11, 12]:
                                cell.font = Font(name='Calibri', size=11, bold=True, color="FF0000")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            else:
                                cell.font = Font(name='Calibri', size=11, bold=False, color="000000")
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # --- SHEET 3: SUMMARY ANALYTICS ---
                    ws_summary = workbook["Summary_Analytics"]
                    ws_summary.column_dimensions['A'].width = 50
                    ws_summary.column_dimensions['B'].width = 30
                    ws_summary.column_dimensions['C'].width = 50
                    
                    for row_num, row in enumerate(ws_summary.iter_rows()):
                        row_idx = row_num + 1
                        is_gray = (row_idx % 2 == 0)
                        
                        for cell in row:
                            if cell.value is not None:
                                cell.border = thin_border
                                cell.alignment = Alignment(wrap_text=True, vertical="center")
                                
                                if row_idx == 1 or row_idx == start_row_for_lists + 1:
                                    cell.fill = header_fill
                                    cell.font = font_12b_white
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                else:
                                    cell.fill = alt_row_fill if is_gray else white_fill
                                    cell.font = Font(name='Calibri', size=11, color="000000")

                print_excel_data = output_print.getvalue()
                
                # ==========================================
                # DISPLAY METRICS & DOWNLOAD BUTTONS
                # ==========================================
                st.success("✅ Reports Generated Successfully!")
                st.subheader("📊 Quick Analytics")
                colA, colB, colC = st.columns(3)
                colA.metric("Provided Companies", total_provided)
                colB.metric("Found in Sales", len(found_in_sales))
                colC.metric("Found in Stock", len(found_in_stock))
                
                st.markdown("---")
                btn1, btn2 = st.columns(2)
                with btn1:
                    st.download_button(
                        "📥 Download Raw Excel Report", 
                        data=raw_excel_data, 
                        file_name="CleanReport_Raw_Data.xlsx", 
                        use_container_width=True
                    )
                with btn2:
                    st.download_button(
                        "🖨️ Download Printable Clean Report", 
                        data=print_excel_data, 
                        file_name="CleanReport_Printable.xlsx", 
                        use_container_width=True,
                        type="primary"
                    )
                
            except Exception as e:
                st.error(f"Error during processing: {e}")
    else:
        st.warning("⚠️ Please upload both Master files and provide target companies!")